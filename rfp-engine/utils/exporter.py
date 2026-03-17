import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

class SovereignExporter:
    """
    Sovereign Intelligence Terminal: Excel Export Engine
    Generates industrial-grade RTM (Requirements Traceability Matrix).
    """

    def __init__(self):
        # Machined Aesthetic: Dark Grey Header, White Text
        self.header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, name="Arial")
        self.cell_font = Font(name="Arial", size=10)
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )

    def generate_rtm(self, compliance_data):
        """
        Generates an Excel RTM from compliance JSON data.
        Returns a BytesIO stream.
        """
        # compliance_data expected: list of dicts with [id, source, text, status, remediation]
        df = pd.DataFrame(compliance_data)
        
        # Ensure correct column order for GovCon RTM
        cols = ['id', 'source', 'text', 'status', 'remediation', 'proposal_mapping']
        for col in cols:
            if col not in df.columns:
                df[col] = "" # Placeholder for empty columns
        
        df = df[cols]
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Compliance Matrix')
            
            # Access openpyxl objects for styling
            workbook = writer.book
            worksheet = writer.sheets['Compliance Matrix']
            
            # Style Headers
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self.thin_border
            
            # Style Body & Column Widths
            dims = {
                'A': 10,  # ID
                'B': 15,  # Source
                'C': 60,  # Text
                'D': 15,  # Status
                'E': 50,  # Remediation
                'F': 30   # Mapping
            }
            
            for col, width in dims.items():
                worksheet.column_dimensions[col].width = width
                
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = self.cell_font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Freeze Top Row
            worksheet.freeze_panes = 'A2'
            
        return output.getvalue()

# Example usage (Stateless stream)
# exporter = SovereignExporter()
# binary_data = exporter.generate_rtm(data)
